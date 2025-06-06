# scheduler_module.py
# -*- coding: utf-8 -*-
import psycopg2
import psycopg2.extras
import datetime
import math
import random
import copy
import pandas as pd
from collections import defaultdict, namedtuple
import re
import io

# --- 检查 openpyxl 库 ---
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font, Border, Side

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    # print("警告：未找到 'openpyxl' 库，将无法导出 Excel 课表。请运行 'pip install openpyxl' 安装。") # 在app.py中处理

# ==================================
# 2. 数据模型 (保持不变)
# ==================================
Semester = namedtuple('Semester', ['id', 'name', 'start_date', 'end_date', 'total_weeks'])
Major = namedtuple('Major', ['id', 'name'])
Teacher = namedtuple('Teacher', ['id', 'user_id', 'name'])
Classroom = namedtuple('Classroom', ['id', 'name', 'capacity', 'type'])
Course = namedtuple('Course', ['id', 'name', 'total_sessions', 'course_type'])
TimeSlot = namedtuple('TimeSlot', ['id', 'day_of_week', 'period', 'start_time', 'end_time'])
CourseAssignment = namedtuple('CourseAssignment',
                              ['id', 'major_id', 'course_id', 'teacher_id', 'semester_id', 'is_core_course',
                               'expected_students'])
TimetableEntry = namedtuple('TimetableEntry',
                            ['id', 'semester_id', 'major_id', 'course_id', 'teacher_id', 'classroom_id', 'timeslot_id',
                             'week_number', 'assignment_id'])
TeacherPreference = namedtuple('TeacherPreference', ['id', 'teacher_id', 'semester_id', 'timeslot_id', 'preference_type', 'status', 'reason'])

# ==================================
# 3. 数据加载函数 (保持不变)
# ==================================
def load_data_from_db(get_connection_func):
    """从数据库加载所有基础数据，包括教师偏好"""
    print("SCHEDULER: 开始从数据库加载数据...")
    all_data = {}
    conn = None
    cur = None # Define cur outside try
    try:
        conn = get_connection_func()  # 使用传入的函数获取连接
        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

        # 加载学期 (计算 total_weeks)
        cur.execute("SELECT id, name, start_date, end_date FROM semesters")
        raw_semesters = cur.fetchall()
        all_data['semesters'] = {}
        # print("  - 正在加载学期信息并计算总周数...")
        for row in raw_semesters:
            start_date = row['start_date']
            end_date = row['end_date']
            calculated_weeks = 0
            if isinstance(start_date, datetime.date) and isinstance(end_date, datetime.date):
                if end_date >= start_date:
                    delta_days = (end_date - start_date).days + 1
                    calculated_weeks = math.ceil(delta_days / 7)
            all_data['semesters'][row['id']] = Semester(
                id=row['id'], name=row['name'], start_date=start_date,
                end_date=end_date, total_weeks=calculated_weeks
            )
        # print(f"  - 加载并处理了 {len(all_data['semesters'])} 个学期信息")

        # 加载专业
        cur.execute("SELECT id, name FROM majors")
        all_data['majors'] = {row['id']: Major(**row) for row in cur.fetchall()}
        # print(f"  - 加载了 {len(all_data['majors'])} 个专业信息")

        # 加载教师
        cur.execute("SELECT id, username FROM users")
        user_id_to_name = {row['id']: row['username'] for row in cur.fetchall()}
        cur.execute("SELECT id, user_id FROM teachers")
        raw_teachers = cur.fetchall()
        all_data['teachers'] = {}
        for row in raw_teachers:
            teacher_id = row['id']
            user_id = row['user_id']
            teacher_name = user_id_to_name.get(user_id, f"未知用户(ID:{user_id})")
            all_data['teachers'][teacher_id] = Teacher(id=teacher_id, user_id=user_id, name=teacher_name)
        # print(f"  - 加载并处理了 {len(all_data['teachers'])} 个教师信息")

        # 加载教室
        cur.execute("SELECT id, building, room_number, capacity, room_type FROM classrooms")
        all_data['classrooms'] = {}
        for row in cur.fetchall():
            classroom_id = row['id']
            building_name = row['building'] if row['building'] else '未知楼'
            room_num = row['room_number'] if row['room_number'] else '未知号'
            classroom_name = f"{building_name}-{room_num}"
            all_data['classrooms'][classroom_id] = Classroom(
                id=classroom_id, name=classroom_name, capacity=row['capacity'], type=row['room_type']
            )
        # print(f"  - 加载了 {len(all_data['classrooms'])} 个教室信息")

        # 加载课程
        cur.execute("SELECT id, name, total_sessions, course_type FROM courses")
        all_data['courses'] = {row['id']: Course(**row) for row in cur.fetchall()}
        # print(f"  - 加载了 {len(all_data['courses'])} 个课程信息")

        # 加载时间段
        cur.execute("SELECT id, day_of_week, period, start_time, end_time FROM time_slots ORDER BY day_of_week, period")
        all_data['timeslots'] = {row['id']: TimeSlot(**row) for row in cur.fetchall()}
        all_data['timeslot_lookup'] = {(ts.day_of_week, ts.period): ts.id for ts in all_data['timeslots'].values()}
        # print(f"  - 加载了 {len(all_data['timeslots'])} 个时间段信息")

        # 加载教学任务
        cur.execute("""
            SELECT id, major_id, course_id, teacher_id, semester_id, is_core_course, expected_students
            FROM course_assignments
        """)
        all_data['course_assignments'] = {row['id']: CourseAssignment(**row) for row in cur.fetchall()}
        # print(f"  - 加载了 {len(all_data['course_assignments'])} 个教学任务")

        # --- 新增：加载教师排课偏好 ---
        cur.execute("""
            SELECT id, teacher_id, semester_id, timeslot_id, preference_type, status, reason
            FROM teacher_scheduling_preferences
            WHERE preference_type = 'avoid'
        """)
        # Store approved 'avoid' preferences in a set for quick lookup: (teacher_id, timeslot_id, semester_id)
        all_data['approved_avoid_preferences'] = set()
        raw_preferences = cur.fetchall()
        for row in raw_preferences:
             all_data['approved_avoid_preferences'].add((row['teacher_id'], row['timeslot_id'], row['semester_id']))
        # print(f"  - 加载并过滤了 {len(raw_preferences)} 条教师偏好记录 (其中 {len(all_data['approved_avoid_preferences'])} 条已批准的'避免'偏好将作为约束)")
        # --- 新增结束 ---

        cur.close()
        print("SCHEDULER: 数据加载成功!")
        return all_data

    except psycopg2.Error as e:
        print(f"SCHEDULER: 数据库连接或查询错误: {e}")
        # import traceback; traceback.print_exc() # For more detailed server-side logs if needed
        raise  # Re-raise the exception to be caught by the Flask route
    except Exception as e:
        print(f"SCHEDULER: 数据加载过程中发生未知错误: {e}")
        # import traceback; traceback.print_exc()
        raise # Re-raise
    finally:
        if cur: cur.close()
        if conn: conn.close()

# ==================================
# 4. 辅助函数 (保持不变)
# ==================================
def find_timeslot_id(day_str, period_num, all_data):
    return all_data['timeslot_lookup'].get((day_str, period_num))

def check_constraints(timetable_state, assignment, week, timeslot_id, classroom_id, all_data):
    teacher_id = assignment.teacher_id
    major_id = assignment.major_id
    semester_id = assignment.semester_id # 获取学期 ID

    # --- 新增：检查教师的“避免安排”偏好 ---
    if (teacher_id, timeslot_id, semester_id) in all_data.get('approved_avoid_preferences', set()):
        # print(f"[CONFLICT] Teacher {teacher_id} has 'avoid' preference for timeslot {timeslot_id} in semester {semester_id}")
        return False, "教师偏好 (避免安排)"
    # --- 新增结束 ---

    # 检查全局状态中教师、教室、专业是否已被占用
    if (teacher_id, week, timeslot_id) in timetable_state['teacher_schedule']:
        # print(f"[CONFLICT] Teacher {teacher_id} busy week {week} slot {timeslot_id}")
        return False, "教师冲突 (已安排其它课程)"
    if (classroom_id, week, timeslot_id) in timetable_state['classroom_schedule']:
        # print(f"[CONFLICT] Classroom {classroom_id} busy week {week} slot {timeslot_id}")
        return False, "教室冲突 (已被占用)"
    if (major_id, week, timeslot_id) in timetable_state['major_schedule']:
        # print(f"[CONFLICT] Major {major_id} busy week {week} slot {timeslot_id}")
        return False, "专业冲突 (已安排其它课程)"

    return True, None # 没有冲突

def find_available_classroom(timetable_state, assignment, week, timeslot_id, all_data):
    required_capacity = assignment.expected_students
    course = all_data['courses'].get(assignment.course_id)
    is_lab_course = course and course.course_type == '实验课'
    preferred_type_available, other_type_available = [], []

    # 确定在当前周次和时间段已经被占用的教室集合，检查全局状态
    busy_classrooms = {cid for (cid, w, tid) in timetable_state['classroom_schedule'] if
                       w == week and tid == timeslot_id}

    for classroom_id, classroom in all_data['classrooms'].items():
        # 检查教室是否在全局状态中已被占用
        if classroom_id in busy_classrooms: continue
        if classroom.capacity < required_capacity: continue
        classroom_type_match = (is_lab_course and classroom.type == '实验室') or \
                               (not is_lab_course and classroom.type == '普通教室')
        if classroom_type_match:
            preferred_type_available.append(classroom_id)
        else:
            other_type_available.append(classroom_id)

    if preferred_type_available: return random.choice(preferred_type_available)
    if other_type_available: return random.choice(other_type_available)
    return None

# ==================================
# 5. 自动生成初始模板函数 (保持不变)
# ==================================
def generate_initial_template(assignments_dict, all_data):
    # print("SCHEDULER:   正在根据可用任务自动生成初始周模板...")
    day_order = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']
    template_structure = [
        ('周一', 1, '理论课', 'MonWed1'), ('周一', 2, '理论课', 'MonWed2'), ('周一', 3, None, None),
        ('周一', 4, None, None),
        ('周二', 1, '理论课', 'TueThu1'), ('周二', 2, '理论课', 'TueThu2'), ('周二', 3, None, None),
        ('周二', 4, None, None),
        ('周三', 1, '理论课', 'MonWed1'), ('周三', 2, '理论课', 'MonWed2'), ('周三', 3, None, None),
        ('周三', 4, None, None),
        ('周四', 1, '理论课', 'TueThu1'), ('周四', 2, '理论课', 'TueThu2'), ('周四', 3, None, None),
        ('周四', 4, None, None),
        ('周五', 1, '实验课', 'FriLab1'), ('周五', 2, '实验课', 'FriLab2'), ('周五', 3, None, None),
        ('周五', 4, None, None),
        ('周六', 1, None, None), ('周六', 2, None, None), ('周六', 3, None, None), ('周六', 4, None, None),
        ('周日', 1, None, None), ('周日', 2, None, None), ('周日', 3, None, None), ('周日', 4, None, None),
    ] # 包含周六周日，即使默认不排课，但时间段可能存在

    repetition_map = defaultdict(list)
    slot_type_map = {}
    initial_template = {}
    all_slots_in_template = set()

    db_timeslots_raw = sorted(all_data['timeslots'].values(), key=lambda t: (t.day_of_week, t.period))
    db_template_structure = [(ts.day_of_week, ts.period, None, None) for ts in db_timeslots_raw] # Default type None, no groups
    id_to_day_period = {ts.id: (ts.day_of_week, ts.period) for ts in all_data['timeslots'].values()}
    initial_template_by_id = {ts_id: None for ts_id in all_data['timeslots'].keys()}
    sorted_timeslot_ids = sorted(all_data['timeslots'].keys(),
                                 key=lambda ts_id: (day_order.index(all_data['timeslots'][ts_id].day_of_week) if all_data['timeslots'][ts_id].day_of_week in day_order else 99,
                                                    all_data['timeslots'][ts_id].period))

    template_structure_simplified = [
        ('周一', 1), ('周一', 2), ('周一', 3), ('周一', 4),
        ('周二', 1), ('周二', 2), ('周二', 3), ('周二', 4),
        ('周三', 1), ('周三', 2), ('周三', 3), ('周三', 4),
        ('周四', 1), ('周四', 2), ('周四', 3), ('周四', 4),
        ('周五', 1), ('周五', 2), ('周五', 3), ('周五', 4),
        ('周六', 1), ('周六', 2), ('周六', 3), ('周六', 4),
        ('周日', 1), ('周日', 2), ('周日', 3), ('周日', 4),
    ]
    template_slots_from_db = [
        (day, period) for day, period in template_structure_simplified
        if (day, period) in all_data['timeslot_lookup']
    ]
    day_order_indices = {day: idx for idx, day in enumerate(day_order)} # Precompute indices for sorting
    sorted_template_slots_dp = sorted(
        template_slots_from_db,
        key=lambda x: (day_order_indices.get(x[0], 99), x[1]) # Use precomputed indices
    )

    initial_template_fill = {} # Maps (day, period) to assignment_id
    all_template_slots_dp = set(sorted_template_slots_dp)

    if not assignments_dict: return {}, [] # Return empty template and pool

    def get_priority(assign_id, assign):
        course = all_data['courses'].get(assign.course_id)
        return (assign.is_core_course, -(course.total_sessions if course else 0), random.random())

    all_assignments_sorted = sorted(
         ((get_priority(assign_id, assign), assign_id) for assign_id, assign in assignments_dict.items()),
         reverse=True
    )
    assignment_pool_ids = [assign_id for _, assign_id in all_assignments_sorted]

    used_assignment_ids = set()
    template_slot_index = 0

    while assignment_pool_ids and template_slot_index < len(sorted_template_slots_dp):
        assign_id_to_fill = assignment_pool_ids.pop(0) # Take the highest priority assignment
        if assign_id_to_fill in used_assignment_ids: continue
        assign = assignments_dict.get(assign_id_to_fill)
        if not assign: continue
        course = all_data['courses'].get(assign.course_id)
        if not course or course.total_sessions <= 0: continue

        current_slot_dp = sorted_template_slots_dp[template_slot_index]

        initial_template_fill[current_slot_dp] = assign_id_to_fill
        used_assignment_ids.add(assign_id_to_fill)
        template_slot_index += 1

    unscheduled_pool_ids = [
        assign_id for assign_id in assignments_dict
        if assign_id not in used_assignment_ids
        and all_data['courses'].get(assignments_dict[assign_id].course_id, Course(None, None, 0, None)).total_sessions > 0
    ]
    # print(f"SCHEDULER:   自动生成模板完成。选入 {len(used_assignment_ids)} 个任务到模板。剩余 {len(unscheduled_pool_ids)} 个任务进入替换池。")

    return initial_template_fill, unscheduled_pool_ids


# ==================================
# 6. 基于模板的排课执行函数 (**核心修改**)
# ==================================
def schedule_with_generated_template(assignments_for_major, current_semester, current_major, all_data, initial_template_dp, # initial_template is (day, period) map
                                     unscheduled_pool_ids, global_timetable_state):
    print(f"\nSCHEDULER: ===== 开始为专业 '{current_major.name}' 排课 (学期: {current_semester.name}, {current_semester.total_weeks} 周) - 采用固定周模板策略 =====")
    total_weeks = current_semester.total_weeks
    if not total_weeks or total_weeks <= 0:
        unscheduled_details_on_error = []
        for assign_id, assign in assignments_for_major.items():
            course = all_data['courses'].get(assign.course_id)
            teacher = all_data['teachers'].get(assign.teacher_id)
            unscheduled_details_on_error.append({
                'assignment_id': assign_id,
                'course_name': course.name if course else '未知课程',
                'teacher_name': teacher.name if teacher else '未知教师',
                'remaining_sessions': course.total_sessions if course else 0
            })
        return {'schedule': [], 'unscheduled_details': unscheduled_details_on_error, 'conflicts': []}

    assignment_sessions_remaining = {}
    for assign_id, assign in assignments_for_major.items():
        course = all_data['courses'].get(assign.course_id)
        assignment_sessions_remaining[assign_id] = course.total_sessions if course else 0

    final_schedule = []
    conflicts_log_week1 = [] # 只记录第一周生成模板时的冲突
    week1_schedule_entries = [] # 存储第一周成功排课的条目
    week1_fixed_template = {} # 新增：存储第一周成功排课的固定模板 {timeslot_id: (assignment_id, classroom_id)}

    # --- 时间段和动态池准备 (用于第一周) ---
    day_order = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']
    sorted_timeslot_ids_by_dp = sorted(
        all_data['timeslots'].keys(),
        key=lambda ts_id: (day_order.index(all_data['timeslots'][ts_id].day_of_week) if all_data['timeslots'][ts_id].day_of_week in day_order else 99,
                           all_data['timeslots'][ts_id].period)
    )
    timeslot_id_to_dp = {ts.id: (ts.day_of_week, ts.period) for ts in all_data['timeslots'].values()}

    # 动态未排池 (仅用于第一周)
    dynamic_unscheduled_assignments_week1 = list(unscheduled_pool_ids) # Start with assignments not in initial template
    # Add assignments from the template if they need sessions, shuffle later
    for dp, assign_id in initial_template_dp.items():
        if assign_id not in dynamic_unscheduled_assignments_week1 and assignment_sessions_remaining.get(assign_id, 0) > 0:
            dynamic_unscheduled_assignments_week1.append(assign_id)
    random.shuffle(dynamic_unscheduled_assignments_week1)


    # --- Phase 1: 排列第一周 (Week 1) 并生成固定模板 ---
    print(f"SCHEDULER:   - 正在排列第 1 周并生成固定模板...")
    week = 1
    assignments_tried_this_week = set() # 避免重复尝试

    for timeslot_id in sorted_timeslot_ids_by_dp:
        day_str, period_num = timeslot_id_to_dp.get(timeslot_id, (None, None))
        if day_str is None: continue

        assignment_to_attempt_id = None
        assignment_source = None # 'template' or 'pool'

        # 1. 尝试初始模板建议
        suggested_assign_id = initial_template_dp.get((day_str, period_num))
        if suggested_assign_id is not None and \
           assignment_sessions_remaining.get(suggested_assign_id, 0) > 0 and \
           suggested_assign_id not in assignments_tried_this_week:
            assignment_to_attempt_id = suggested_assign_id
            assignment_source = 'template'
            # print(f"  W1, {day_str}-{period_num}: Trying template task {suggested_assign_id}")

        # 2. 如果模板建议不行或已尝试，尝试从未排池选择
        if assignment_to_attempt_id is None:
            for assign_id_from_pool in list(dynamic_unscheduled_assignments_week1): # Iterate copy
                if assignment_sessions_remaining.get(assign_id_from_pool, 0) > 0 and \
                   assign_id_from_pool not in assignments_tried_this_week:
                    assign_from_pool = assignments_for_major.get(assign_id_from_pool)
                    # 快速检查教师偏好
                    if assign_from_pool and (assign_from_pool.teacher_id, timeslot_id, assign_from_pool.semester_id) in all_data.get('approved_avoid_preferences', set()):
                        continue # Skip due to preference

                    assignment_to_attempt_id = assign_id_from_pool
                    assignment_source = 'pool'
                    dynamic_unscheduled_assignments_week1.remove(assign_id_from_pool) # Take from pool
                    # print(f"  W1, {day_str}-{period_num}: Trying pool task {assign_id_from_pool}")
                    break # Found one from pool

        # 3. 尝试安排选定的任务
        if assignment_to_attempt_id is not None:
            assignments_tried_this_week.add(assignment_to_attempt_id) # Mark as tried this week
            assignment = assignments_for_major.get(assignment_to_attempt_id)
            if not assignment: continue

            # 找教室
            suitable_classroom_id = find_available_classroom(global_timetable_state, assignment, week, timeslot_id, all_data)

            if suitable_classroom_id:
                # 检查约束
                is_possible, conflict_reason = check_constraints(global_timetable_state, assignment, week, timeslot_id, suitable_classroom_id, all_data)

                if is_possible:
                    # 成功安排第一周！
                    entry = TimetableEntry(None, current_semester.id, assignment.major_id, assignment.course_id,
                                           assignment.teacher_id, suitable_classroom_id, timeslot_id, week,
                                           assignment_to_attempt_id)
                    final_schedule.append(entry) # 加入最终列表
                    week1_schedule_entries.append(entry) # 加入第一周列表，用于建模板
                    # 构建固定模板
                    week1_fixed_template[timeslot_id] = (assignment_to_attempt_id, suitable_classroom_id)

                    # 更新全局状态
                    global_timetable_state['teacher_schedule'].add((assignment.teacher_id, week, timeslot_id))
                    global_timetable_state['classroom_schedule'].add((suitable_classroom_id, week, timeslot_id))
                    global_timetable_state['major_schedule'].add((assignment.major_id, week, timeslot_id))

                    # 减少剩余课时
                    assignment_sessions_remaining[assignment_to_attempt_id] -= 1
                    # print(f"  SUCCESS W1: {day_str}-{period_num} assigned {assignment_to_attempt_id} in C{suitable_classroom_id}. Remaining: {assignment_sessions_remaining[assignment_to_attempt_id]}")

                    # 如果任务完成，从动态池移除 (如果它在里面)
                    if assignment_sessions_remaining[assignment_to_attempt_id] == 0:
                        if assignment_to_attempt_id in dynamic_unscheduled_assignments_week1:
                            try: dynamic_unscheduled_assignments_week1.remove(assignment_to_attempt_id)
                            except ValueError: pass

                else:
                    # 第一周约束冲突，记录
                    conflicts_log_week1.append(
                        {'major_id': assignment.major_id, 'week': week, 'day': day_str, 'period': period_num,
                         'assignment_id': assignment_to_attempt_id, 'reason': f"W1约束冲突: {conflict_reason}"})
                    # print(f"  CONFLICT W1: {day_str}-{period_num} attempt {assignment_to_attempt_id} failed: {conflict_reason}")
                    # 如果是从池里拿出来的，放回去
                    if assignment_source == 'pool':
                        if assignment_sessions_remaining.get(assignment_to_attempt_id, 0) > 0: # 只有还有课时才放回
                           dynamic_unscheduled_assignments_week1.append(assignment_to_attempt_id)
                           random.shuffle(dynamic_unscheduled_assignments_week1)

            else:
                # 第一周找不到教室，记录
                conflicts_log_week1.append(
                    {'major_id': assignment.major_id, 'week': week, 'day': day_str, 'period': period_num,
                     'assignment_id': assignment_to_attempt_id,
                     'reason': f"W1找不到容量({assignment.expected_students})教室"})
                # print(f"  NOCLASSROOM W1: {day_str}-{period_num} attempt {assignment_to_attempt_id} failed: No suitable classroom.")
                # 如果是从池里拿出来的，放回去
                if assignment_source == 'pool':
                     if assignment_sessions_remaining.get(assignment_to_attempt_id, 0) > 0:
                        dynamic_unscheduled_assignments_week1.append(assignment_to_attempt_id)
                        random.shuffle(dynamic_unscheduled_assignments_week1)
        # else:
            # print(f"  W1, {day_str}-{period_num}: No suitable assignment found or all tried/finished.")


    # --- Phase 2: 复制固定模板到后续周次 (Weeks 2 to N) ---
    print(f"SCHEDULER:   - 第 1 周模板生成完毕 (排入 {len(week1_fixed_template)} 个时段)。开始复制到后续周...")
    if not week1_fixed_template:
         print(f"SCHEDULER:   - 警告：专业 '{current_major.name}' 未能在第 1 周排入任何课程，无法生成固定模板。")
    else:
        for week in range(2, total_weeks + 1):
            # print(f"SCHEDULER:     - 正在复制模板到第 {week} 周...")
            for timeslot_id, (assignment_id, classroom_id) in week1_fixed_template.items():
                # 检查模板中的任务是否还有剩余课时
                if assignment_sessions_remaining.get(assignment_id, 0) > 0:
                    assignment = assignments_for_major.get(assignment_id)
                    if not assignment: continue # 任务数据丢失？

                    # 检查全局状态，看这个资源是否已被 *其他专业* 在本周本时段占用 (理论上不应检查自身冲突，因为是复制)
                    # 注意：这里简化处理，不完全重新检查所有约束，主要防止与其他专业冲突
                    teacher_busy = (assignment.teacher_id, week, timeslot_id) in global_timetable_state['teacher_schedule']
                    classroom_busy = (classroom_id, week, timeslot_id) in global_timetable_state['classroom_schedule']
                    major_busy = (assignment.major_id, week, timeslot_id) in global_timetable_state['major_schedule'] # 理论上不应发生

                    if teacher_busy or classroom_busy or major_busy:
                         # 记录潜在的周间冲突（通常是由于其他专业抢占了资源）
                         reason = []
                         if teacher_busy: reason.append("教师已被占用")
                         if classroom_busy: reason.append("教室已被占用")
                         if major_busy: reason.append("专业时段已被占用")
                         # 这里选择记录冲突并跳过，还是强制覆盖？目前选择跳过以避免硬冲突
                         ts_info = all_data['timeslots'].get(timeslot_id)
                         day_str = ts_info.day_of_week if ts_info else '?'
                         period_num = ts_info.period if ts_info else '?'
                         conflicts_log_week1.append( # 把周间冲突也加入日志
                             {'major_id': assignment.major_id, 'week': week, 'day': day_str, 'period': period_num,
                              'assignment_id': assignment_id,
                              'reason': f"W{week}模板复制冲突: {', '.join(reason)} (被其他专业占用?)"})
                         # print(f"  SKIP W{week}: Template slot {day_str}-{period_num} for {assignment_id} skipped due to conflict: {', '.join(reason)}")
                         continue # 跳过这个时段的复制

                    # 创建排课条目 (直接使用模板信息)
                    entry = TimetableEntry(None, current_semester.id, assignment.major_id, assignment.course_id,
                                           assignment.teacher_id, classroom_id, timeslot_id, week,
                                           assignment_id)
                    final_schedule.append(entry)

                    # 更新全局状态 (重要！通知其他专业此资源已占用)
                    global_timetable_state['teacher_schedule'].add((assignment.teacher_id, week, timeslot_id))
                    global_timetable_state['classroom_schedule'].add((classroom_id, week, timeslot_id))
                    global_timetable_state['major_schedule'].add((assignment.major_id, week, timeslot_id))

                    # 减少剩余课时
                    assignment_sessions_remaining[assignment_id] -= 1
                    # print(f"  REPLICATED W{week}: Slot {day_str}-{period_num} assigned {assignment_id} in C{classroom_id}. Remaining: {assignment_sessions_remaining[assignment_id]}")

    # --- Final Check: 未完成的任务 ---
    unscheduled_final = []
    for assign_id, remaining in assignment_sessions_remaining.items():
        if remaining > 0:
            assign_obj = assignments_for_major.get(assign_id)
            course_name, teacher_name = '?', '?'
            if assign_obj:
                course = all_data['courses'].get(assign_obj.course_id)
                teacher = all_data['teachers'].get(assign_obj.teacher_id)
                if course: course_name = course.name
                if teacher: teacher_name = teacher.name
            unscheduled_final.append(
                {'assignment_id': assign_id, 'course_name': course_name, 'teacher_name': teacher_name,
                 'remaining_sessions': remaining})

    print(f"SCHEDULER: ===== 专业 '{current_major.name}' 固定模板排课完成。总生成课表条目: {len(final_schedule)}, 第一周冲突记录: {len(conflicts_log_week1)}, 最终未完成任务数: {len(unscheduled_final)} =====")
    # 返回结果，冲突只包含第一周生成模板时的冲突
    return {'schedule': final_schedule, 'unscheduled_details': unscheduled_final, 'conflicts': conflicts_log_week1}


# ==================================
# 7. 导出到 Excel 函数 (保持不变)
# ==================================
def sanitize_sheet_name(name):
    name = re.sub(r'[\\/*?:"<>|\[\]]', '_', name)
    return name[:31]

def format_semester_sheet_for_export(worksheet, total_weeks, periods_count):
    if not OPENPYXL_AVAILABLE: return
    try:
        center_alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        thin_border_side = Side(border_style="thin", color="000000")
        thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side,
                             bottom=thin_border_side)
        header_font = Font(bold=True)

        for col_idx, column_cells in enumerate(worksheet.columns):
            if col_idx < 2: continue # Skip '周数' and '节次' columns for width calculation
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                if cell.value:
                    try:
                        # Consider text wrapping for length calculation
                        lines = str(cell.value).split('\n')
                        # Heuristic: Estimate width based on characters per line
                        # Adjust multiplier based on typical font; 1.8 for CJK, 1 for others might be too simple
                        # Let's try a simpler approach: max length of any line + buffer
                        cell_len = max(len(line) for line in lines) if lines else 0
                        # Give more weight to CJK chars maybe? Or just use a fixed multiplier for max line len?
                        # Example: Max line length * 1.2 + 2? Experiment needed.
                        # Let's stick to a simpler max char count + buffer for now.
                        max_line_len = 0
                        for line in lines:
                             line_len = sum(1.8 if '\u4e00' <= char <= '\u9fff' else 1 for char in line)
                             if line_len > max_line_len:
                                 max_line_len = line_len

                        if max_line_len > max_length: max_length = max_line_len
                    except:
                        pass # Ignore errors during length calculation

            # Set width: baseline + calculated length + some padding
            adjusted_width = max(max_length + 4, 15) # Min width 15, add padding
            worksheet.column_dimensions[column_letter].width = adjusted_width


        worksheet.column_dimensions['A'].width = 8  # 周数
        worksheet.column_dimensions['B'].width = 8  # 节次

        header_rows = 1 # Assuming 1 header row for MultiIndex
        # Apply styles to all cells
        for row in worksheet.iter_rows():
             for cell in row:
                 cell.alignment = center_alignment
                 cell.border = thin_border
                 # Set default row height (might need adjustment based on content)
                 worksheet.row_dimensions[cell.row].height = 45

        # Style the header row specifically (row 1 for MultiIndex level 0)
        # And potentially row 2 for level 1 if MultiIndex headers are split
        # Check how pandas writes MultiIndex headers to be sure
        # Assuming headers are in row 1 and maybe 2
        for col_cell in worksheet[1]: # First row (days)
             col_cell.font = header_font
        if worksheet.max_row >= 2: # If multi-index headers span two rows
             for col_cell in worksheet[2]:
                 col_cell.font = header_font # Apply to second header row too if needed

        # Apply header font to index names (周数, 节次) - these are usually in specific cells
        # Check where pandas places them. Typically A1, B1 or A2, B2 etc.
        try:
            # Assuming index names are at the start of the header rows
             worksheet['A1'].font = header_font # 周数
             worksheet['B1'].font = header_font # 节次
             # If MultiIndex headers shift things, adjust accordingly
        except Exception:
             pass # Ignore if cells don't exist or indexing is different

        # Merge '周数' column cells
        if total_weeks > 0 and periods_count > 0:
            # Need to account for header rows when calculating merge range
            start_header_row = 1 # Row where the actual data starts (after headers)
            # Find the actual first row of data (where index starts)
            # This depends on how pandas writes the multi-index. Let's assume 2 header rows.
            first_data_row = 3
            if worksheet.cell(row=1, column=1).value == "周数": first_data_row = 2 # If index names are on row 1
            if worksheet.cell(row=2, column=1).value == "周数": first_data_row = 3 # If index names are on row 2

            for week in range(1, total_weeks + 1):
                start_row = first_data_row + (week - 1) * periods_count
                end_row = start_row + periods_count - 1

                if start_row <= worksheet.max_row and end_row <= worksheet.max_row and start_row <= end_row:
                    try:
                        # Check if already merged: openpyxl raises ValueError if trying to merge over existing merges
                        is_merged = False
                        cell_coord = f"A{start_row}"
                        for merged_range in worksheet.merged_cells.ranges:
                            if cell_coord in merged_range:
                                is_merged = True
                                break
                        if not is_merged:
                             worksheet.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
                             merged_cell = worksheet.cell(row=start_row, column=1)
                             merged_cell.value = f"第 {week} 周"
                             merged_cell.alignment = center_alignment
                             # merged_cell.font = header_font # Optional: make week number bold
                             merged_cell.border = thin_border # Ensure border is applied to merged cell
                        else:
                             # If already merged, just ensure value and style (might be needed if file is regenerated)
                             merged_cell = worksheet.cell(row=start_row, column=1)
                             if merged_cell.value != f"第 {week} 周": merged_cell.value = f"第 {week} 周"
                             merged_cell.alignment = center_alignment
                             # merged_cell.font = header_font
                             merged_cell.border = thin_border


                    except ValueError as merge_error:
                        # print(f"SCHEDULER: 警告：尝试合并单元格 (周数 {week}, 行 {start_row}-{end_row}) 时出错: {merge_error}. 可能已合并?")
                        # If merge fails, at least set the value in the first cell
                        first_cell = worksheet.cell(row=start_row, column=1)
                        if first_cell.value != f"第 {week} 周": first_cell.value = f"第 {week} 周"
                        first_cell.alignment = center_alignment
                        # first_cell.font = header_font
                        first_cell.border = thin_border

                    except Exception as merge_e:
                         print(f"SCHEDULER: 警告：合并单元格 (周数 {week}) 时发生未知错误: {merge_e}")
                         # Fallback: set value in the first cell
                         first_cell = worksheet.cell(row=start_row, column=1)
                         if first_cell.value != f"第 {week} 周": first_cell.value = f"第 {week} 周"
                         first_cell.alignment = center_alignment
                         # first_cell.font = header_font
                         first_cell.border = thin_border

    except Exception as format_e:
        print(f"SCHEDULER: 警告：在调整工作表 '{worksheet.title}' 格式时发生错误: {format_e}")
        # import traceback; traceback.print_exc() # Debugging help

def generate_excel_report_for_send_file(schedule_entries, all_data, semester,
                                        target_major_id=None, target_teacher_id=None):
    if not OPENPYXL_AVAILABLE:
        raise ImportError("缺少 openpyxl 库，无法导出 Excel 课表。")

    # Handle cases where there's absolutely no data or filter yields nothing
    create_empty_excel = False
    empty_message = "没有排课数据可导出。"
    if not schedule_entries:
        create_empty_excel = True
        if target_major_id:
            major_info = all_data['majors'].get(target_major_id)
            major_name = major_info.name if major_info else f"专业ID_{target_major_id}"
            empty_message = f"专业 '{major_name}' 没有排课数据。"
        elif target_teacher_id:
            teacher_info = all_data['teachers'].get(target_teacher_id)
            teacher_name = teacher_info.name if teacher_info else f"教师ID_{target_teacher_id}"
            empty_message = f"教师 '{teacher_name}' 没有排课数据。"
    elif target_major_id and not any(e.major_id == target_major_id for e in schedule_entries):
         create_empty_excel = True
         major_info = all_data['majors'].get(target_major_id)
         major_name = major_info.name if major_info else f"专业ID_{target_major_id}"
         empty_message = f"专业 '{major_name}' 没有找到符合条件的排课数据。"
    elif target_teacher_id and not any(e.teacher_id == target_teacher_id for e in schedule_entries):
         create_empty_excel = True
         teacher_info = all_data['teachers'].get(target_teacher_id)
         teacher_name = teacher_info.name if teacher_info else f"教师ID_{target_teacher_id}"
         empty_message = f"教师 '{teacher_name}' 没有找到符合条件的排课数据。"


    if create_empty_excel:
        output_buffer_empty = io.BytesIO()
        try:
            with pd.ExcelWriter(output_buffer_empty, engine='openpyxl') as writer_empty:
                df_empty = pd.DataFrame([[empty_message]])
                df_empty.to_excel(writer_empty, sheet_name="无数据", index=False, header=False)
                # Format the single cell? Maybe not necessary.
            output_buffer_empty.seek(0)
        except Exception as empty_excel_e:
            print(f"SCHEDULER: 生成空 Excel 文件时出错: {empty_excel_e}")
            return io.BytesIO() # Return truly empty buffer on error
        return output_buffer_empty


    # print(f"SCHEDULER: 开始生成 Excel 报告...")
    output_buffer = io.BytesIO()
    try:
        with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
            timeslots = sorted(all_data['timeslots'].values(), key=lambda t: (t.day_of_week, t.period))
            day_map = {'周一': 1, '周二': 2, '周三': 3, '周四': 4, '周五': 5, '周六': 6, '周日': 7}
            periods = sorted(list(set(t.period for t in timeslots)))
            days = sorted(list(set(t.day_of_week for t in timeslots)), key=lambda d: day_map.get(d, 8))
            total_weeks = semester.total_weeks

            if total_weeks <= 0:
                print("SCHEDULER: 错误：学期总周数无效，无法生成学期课表。")
                df_error = pd.DataFrame([["学期总周数无效，无法生成课表"]])
                df_error.to_excel(writer, sheet_name="错误", index=False, header=False)
                output_buffer.seek(0)
                return output_buffer

            periods_count = len(periods)
            all_defined_days = sorted(list(set(ts.day_of_week for ts in all_data['timeslots'].values())), key=lambda d: day_map.get(d, 8))
            if len(days) != len(all_defined_days): days = all_defined_days

            # Filter entries based on target_major_id or target_teacher_id *before* processing
            filtered_entries = schedule_entries
            entity_name = None # Store name for sheet naming

            if target_major_id:
                filtered_entries = [e for e in schedule_entries if e.major_id == target_major_id]
                major_info = all_data['majors'].get(target_major_id)
                entity_name = major_info.name if major_info else f"专业ID_{target_major_id}"
            elif target_teacher_id:
                filtered_entries = [e for e in schedule_entries if e.teacher_id == target_teacher_id]
                teacher_info = all_data['teachers'].get(target_teacher_id)
                entity_name = teacher_info.name if teacher_info else f"教师ID_{target_teacher_id}"

            # If after filtering, no entries remain (this check is redundant if handled above, but safe)
            if not filtered_entries and (target_major_id or target_teacher_id):
                 # This case should have been handled by create_empty_excel logic
                 print("SCHEDULER: 逻辑错误 - 筛选后无数据，但未生成空Excel。")
                 return io.BytesIO() # Return empty

            # --- Generate Sheet(s) ---
            if target_major_id:
                # Generate single sheet for the major
                df_major_semester = pd.DataFrame(
                    index=pd.MultiIndex.from_product([range(1, total_weeks + 1), periods], names=['周数', '节次']),
                    columns=days
                ).fillna('')
                for entry in filtered_entries:
                    ts = all_data['timeslots'].get(entry.timeslot_id)
                    if not ts or ts.day_of_week not in days or ts.period not in periods: continue
                    course = all_data['courses'].get(entry.course_id)
                    teacher = all_data['teachers'].get(entry.teacher_id)
                    classroom = all_data['classrooms'].get(entry.classroom_id)
                    cell_text = f"{course.name if course else '?'}\n{teacher.name if teacher else '?'}\n@{classroom.name if classroom else '?'}"
                    if (entry.week_number, ts.period) in df_major_semester.index and ts.day_of_week in df_major_semester.columns:
                         df_major_semester.loc[(entry.week_number, ts.period), ts.day_of_week] = cell_text

                sheet_name = sanitize_sheet_name(f"专业_{entity_name}")
                df_major_semester.to_excel(writer, sheet_name=sheet_name, merge_cells=False)
                worksheet = writer.sheets[sheet_name]
                format_semester_sheet_for_export(worksheet, total_weeks, periods_count)

            elif target_teacher_id:
                # Generate single sheet for the teacher
                df_teacher_semester = pd.DataFrame(
                    index=pd.MultiIndex.from_product([range(1, total_weeks + 1), periods], names=['周数', '节次']),
                    columns=days
                ).fillna('')
                for entry in filtered_entries:
                    ts = all_data['timeslots'].get(entry.timeslot_id)
                    if not ts or ts.day_of_week not in days or ts.period not in periods: continue
                    course = all_data['courses'].get(entry.course_id)
                    major = all_data['majors'].get(entry.major_id)
                    classroom = all_data['classrooms'].get(entry.classroom_id)
                    cell_text = f"{course.name if course else '?'}\n({major.name if major else '?'})\n@{classroom.name if classroom else '?'}"
                    if (entry.week_number, ts.period) in df_teacher_semester.index and ts.day_of_week in df_teacher_semester.columns:
                         df_teacher_semester.loc[(entry.week_number, ts.period), ts.day_of_week] = cell_text

                sheet_name = sanitize_sheet_name(f"教师_{entity_name}")
                df_teacher_semester.to_excel(writer, sheet_name=sheet_name, merge_cells=False)
                worksheet = writer.sheets[sheet_name]
                format_semester_sheet_for_export(worksheet, total_weeks, periods_count)

            else:  # Full semester report (all majors AND all teachers sheets)
                # A. 专业课表 (One sheet per major involved)
                schedule_by_major = defaultdict(list)
                involved_major_ids = sorted(list(set(entry.major_id for entry in filtered_entries)), key=lambda mid: all_data['majors'].get(mid, Major(id=mid, name=f"UnknownMajor{mid}")).name)
                for entry in filtered_entries: schedule_by_major[entry.major_id].append(entry)

                for major_id_iter in involved_major_ids:
                    major_schedule = schedule_by_major[major_id_iter]
                    major_info_iter = all_data['majors'].get(major_id_iter)
                    major_name_iter = major_info_iter.name if major_info_iter else f"专业ID_{major_id_iter}"
                    df_major_semester = pd.DataFrame(
                        index=pd.MultiIndex.from_product([range(1, total_weeks + 1), periods], names=['周数', '节次']),
                        columns=days
                    ).fillna('')
                    for entry in major_schedule:
                        ts = all_data['timeslots'].get(entry.timeslot_id)
                        if not ts or ts.day_of_week not in days or ts.period not in periods: continue
                        course = all_data['courses'].get(entry.course_id)
                        teacher = all_data['teachers'].get(entry.teacher_id)
                        classroom = all_data['classrooms'].get(entry.classroom_id)
                        cell_text = f"{course.name if course else '?'}\n{teacher.name if teacher else '?'}\n@{classroom.name if classroom else '?'}"
                        if (entry.week_number, ts.period) in df_major_semester.index and ts.day_of_week in df_major_semester.columns:
                            df_major_semester.loc[(entry.week_number, ts.period), ts.day_of_week] = cell_text

                    sheet_name_major = sanitize_sheet_name(f"专业_{major_name_iter}")
                    df_major_semester.to_excel(writer, sheet_name=sheet_name_major, merge_cells=False)
                    worksheet_major = writer.sheets[sheet_name_major]
                    format_semester_sheet_for_export(worksheet_major, total_weeks, periods_count)

                # B. 教师课表 (One sheet per teacher involved)
                schedule_by_teacher = defaultdict(list)
                involved_teacher_ids = sorted(list(set(entry.teacher_id for entry in filtered_entries)), key=lambda tid: all_data['teachers'].get(tid, Teacher(id=tid, user_id=None, name=f"UnknownTeacher{tid}")).name)
                for entry in filtered_entries: schedule_by_teacher[entry.teacher_id].append(entry)

                for teacher_id_iter in involved_teacher_ids:
                    teacher_schedule = schedule_by_teacher[teacher_id_iter]
                    teacher_info_iter = all_data['teachers'].get(teacher_id_iter)
                    teacher_name_iter = teacher_info_iter.name if teacher_info_iter else f"教师ID_{teacher_id_iter}"
                    df_teacher_semester = pd.DataFrame(
                        index=pd.MultiIndex.from_product([range(1, total_weeks + 1), periods], names=['周数', '节次']),
                        columns=days
                    ).fillna('')
                    for entry in teacher_schedule:
                        ts = all_data['timeslots'].get(entry.timeslot_id)
                        if not ts or ts.day_of_week not in days or ts.period not in periods: continue
                        course = all_data['courses'].get(entry.course_id)
                        major = all_data['majors'].get(entry.major_id)
                        classroom = all_data['classrooms'].get(entry.classroom_id)
                        cell_text = f"{course.name if course else '?'}\n({major.name if major else '?'})\n@{classroom.name if classroom else '?'}"
                        if (entry.week_number, ts.period) in df_teacher_semester.index and ts.day_of_week in df_teacher_semester.columns:
                            df_teacher_semester.loc[(entry.week_number, ts.period), ts.day_of_week] = cell_text

                    sheet_name_teacher = sanitize_sheet_name(f"教师_{teacher_name_iter}")
                    df_teacher_semester.to_excel(writer, sheet_name=sheet_name_teacher, merge_cells=False)
                    worksheet_teacher = writer.sheets[sheet_name_teacher]
                    format_semester_sheet_for_export(worksheet_teacher, total_weeks, periods_count)

        # End of 'with pd.ExcelWriter' block, file is saved to buffer here
        output_buffer.seek(0)
        # print(f"SCHEDULER: Excel 报告生成完毕。")
        return output_buffer

    except Exception as e:
        print(f"SCHEDULER: 生成 Excel 文件时出错: {e}")
        # import traceback; traceback.print_exc() # For debugging
        # Try to return a buffer with an error message
        error_output = io.BytesIO()
        try:
             with pd.ExcelWriter(error_output, engine='openpyxl') as writer_error:
                 pd.DataFrame([["生成Excel时出错:", str(e)]]).to_excel(writer_error, sheet_name="错误", index=False, header=False)
             error_output.seek(0)
        except Exception as inner_e:
             print(f"SCHEDULER: 写入Excel错误信息时也出错: {inner_e}")
             return io.BytesIO() # Return empty buffer if even error logging fails
        return error_output # Return buffer with error message


# ==================================
# 8. 数据库操作函数 (保持不变)
# ==================================
def clear_db_for_semester(semester_id, get_connection_func):
    conn = None
    cur = None
    try:
        conn = get_connection_func()
        cur = conn.cursor()
        # print(f"SCHEDULER: 正在清空数据库中 学期 ID={semester_id} 的旧排课记录...")
        delete_query = "DELETE FROM timetable_entries WHERE semester_id = %s"
        cur.execute(delete_query, (semester_id,))
        deleted_count = cur.rowcount
        conn.commit()
        cur.close()
        # print(f"SCHEDULER: 成功删除 {deleted_count} 条旧记录。")
        return True, deleted_count
    except psycopg2.Error as e:
        print(f"SCHEDULER: 清空学期 {semester_id} 的数据库记录时出错: {e}")
        if conn: conn.rollback()
        # import traceback; traceback.print_exc()
        raise  # Re-raise
    except Exception as e:
        print(f"SCHEDULER: 清空记录过程中发生未知错误: {e}")
        if conn: conn.rollback()
        raise
    finally:
        if cur: cur.close()
        if conn: conn.close()

def save_schedule_to_db(schedule_entries, get_connection_func):
    if not schedule_entries:
        # print("SCHEDULER: 没有排课条目需要保存。")
        return 0
    conn_save = None
    cur_save = None
    inserted_count = 0
    try:
        conn_save = get_connection_func()
        cur_save = conn_save.cursor()

        insert_query = """
            INSERT INTO timetable_entries
            (semester_id, major_id, course_id, teacher_id, classroom_id, timeslot_id, week_number, assignment_id)
            VALUES %s
        """

        data_to_insert = [
            (e.semester_id, e.major_id, e.course_id, e.teacher_id, e.classroom_id, e.timeslot_id, e.week_number,
             e.assignment_id)
            for e in schedule_entries
        ]

        from psycopg2.extras import execute_values
        execute_values(cur_save, insert_query, data_to_insert, page_size=max(100, len(data_to_insert)//10))

        inserted_count = len(data_to_insert) # Assume all attempted rows are inserted if no error
        conn_save.commit()

        cur_save.close()
        # print(f"SCHEDULER: 成功保存 {inserted_count} 条排课记录到数据库。")
        return inserted_count
    except psycopg2.Error as e:
        print(f"SCHEDULER: 保存排课结果到数据库时出错: {e}")
        if conn_save: conn_save.rollback()
        # import traceback; traceback.print_exc()
        raise  # Re-raise
    except Exception as e:
         print(f"SCHEDULER: 保存记录过程中发生未知错误: {e}")
         if conn_save: conn_save.rollback()
         raise
    finally:
        if cur_save: cur_save.close()
        if conn_save: conn_save.close()


# ==================================
# 9. 主排课流程函数 (保持不变)
# ==================================
from collections import defaultdict
import random
# Assume necessary classes (Course, Major, etc.) and functions are defined elsewhere and correctly imported.
# Assume get_connection_func returns a standard DB-API 2 connection object.

def run_full_scheduling_process(target_semester_id, get_connection_func):
    """
    主排课流程函数，被 Flask API 调用。
    返回一个包含排课结果摘要的字典。
    在排课完成后（无论成功或失败）尝试更新所有教师偏好状态。
    """
    print(f"SCHEDULER: 开始执行学期 ID {target_semester_id} 的自动排课程序...")
    summary = {
        "status": "failure",
        "message": "",
        "processed_majors": 0,
        "total_scheduled_entries": 0,
        "total_conflicts": 0, # Note: Conflicts now primarily reflect Week 1 issues or inter-major conflicts during replication
        "total_uncompleted_tasks": 0,
        "db_records_cleared": 0,
        "db_records_saved": 0,
        "details": []  # For per-major messages or errors
    }

    all_data = None
    all_assignments_in_semester = defaultdict(dict)

    try:
        all_data = load_data_from_db(get_connection_func)
        if not all_data:
            summary["message"] = "数据加载失败。"
            return summary # Finally block will still run

        current_semester = all_data['semesters'].get(target_semester_id)
        if not current_semester:
            summary["message"] = f"未找到 ID 为 {target_semester_id} 的学期信息。"
            return summary # Finally block will still run
        if current_semester.total_weeks <= 0:
             summary["message"] = f"目标学期 '{current_semester.name}' (ID: {target_semester_id}) 总周数 ({current_semester.total_weeks}) 无效。"
             return summary # Finally block will still run

        majors_in_semester = set()
        for assign_id, assign in all_data['course_assignments'].items():
            if assign.semester_id == target_semester_id:
                majors_in_semester.add(assign.major_id)
                all_assignments_in_semester[assign.major_id][assign_id] = assign

        if not majors_in_semester:
            summary["message"] = f"学期 '{current_semester.name}' (ID: {target_semester_id}) 中未找到任何专业的教学任务。"
            summary["status"] = "success_no_tasks"
            return summary # Finally block will still run

        clear_success, cleared_count = clear_db_for_semester(target_semester_id, get_connection_func)
        summary["db_records_cleared"] = cleared_count
        if not clear_success:
             summary["message"] = "清空旧排课记录失败。"
             summary["status"] = "error"
             return summary # Finally block will still run

        all_final_schedule_entries_for_semester = []
        master_global_timetable_state = {'teacher_schedule': set(), 'classroom_schedule': set(), 'major_schedule': set()}

        # Define a safe sort key function
        def get_major_sort_key(major_id):
            major = all_data['majors'].get(major_id)
            return major.name if major else f"未知专业ID_{major_id}"

        sorted_major_ids = sorted(list(majors_in_semester), key=get_major_sort_key)

        for major_id in sorted_major_ids:
            current_major = all_data['majors'].get(major_id)
            assignments_for_this_major = all_assignments_in_semester.get(major_id, {})
            major_name = current_major.name if current_major else f"未知专业ID_{major_id}"

            major_detail_msg = f"专业 '{major_name}' (ID: {major_id}): "
            if not assignments_for_this_major:
                major_detail_msg += "没有教学任务，跳过。"
                summary["details"].append(major_detail_msg)
                continue

            initial_template_dp, unscheduled_pool = generate_initial_template(assignments_for_this_major, all_data)
            major_obj_for_scheduling = current_major if current_major else type('MajorDummy', (object,), {'id': major_id, 'name': major_name})()

            # Call the MODIFIED scheduling function
            schedule_result_obj = schedule_with_generated_template(
                assignments_for_this_major, current_semester,
                major_obj_for_scheduling,
                all_data, initial_template_dp, unscheduled_pool,
                master_global_timetable_state  # Pass and update global state
            )

            major_schedule = schedule_result_obj.get('schedule', [])
            all_final_schedule_entries_for_semester.extend(major_schedule)
            # Global state is updated inside the function now

            num_scheduled_major = len(major_schedule)
            major_conflicts_log = schedule_result_obj.get('conflicts', []) # Now reflects W1/replication conflicts
            num_conflicts_major = len(major_conflicts_log)
            major_unscheduled_details = schedule_result_obj.get('unscheduled_details', [])
            num_uncompleted_major = len(major_unscheduled_details)

            summary["processed_majors"] += 1
            summary["total_scheduled_entries"] += num_scheduled_major
            summary["total_conflicts"] += num_conflicts_major
            summary["total_uncompleted_tasks"] += num_uncompleted_major

            major_detail_msg += f"生成课表 {num_scheduled_major}条, 记录冲突 {num_conflicts_major}次。"
            if num_uncompleted_major > 0:
                 major_detail_msg += f" 未完成任务 {num_uncompleted_major}个。"
            summary["details"].append(major_detail_msg)

        if all_final_schedule_entries_for_semester:
            saved_count_total = save_schedule_to_db(all_final_schedule_entries_for_semester, get_connection_func)
            summary["db_records_saved"] = saved_count_total

        summary["status"] = "success"
        summary["message"] = f"学期 {target_semester_id} 排课完成 (采用固定周模板策略)。"
        if summary["total_conflicts"] > 0:
            summary["message"] += f" 总记录冲突: {summary['total_conflicts']}次。"

    except Exception as e:
        print(f"SCHEDULER: 排课主流程发生严重错误: {e}")
        import traceback; traceback.print_exc() # Keep traceback for debugging errors
        summary["message"] = f"排课过程中发生错误: {str(e)}"
        summary["status"] = "error"

    finally:
        # --- START: Update teacher preference status ---
        print("SCHEDULER: 排课流程结束，尝试更新所有教师偏好状态...")
        update_conn = None
        update_cursor = None
        try:
            update_conn = get_connection_func()
            update_cursor = update_conn.cursor()
            # --- Update ALL preferences status ---
            # IMPORTANT: Define the status value used in your DB
            new_status_value = "applied" # Example: use 'applied', 'processed', etc.
            # Consider if you only want to update preferences for the processed semester
            # or only those that were initially 'pending' etc. Add WHERE clause if needed.
            # Example: WHERE semester_id = %s AND status = 'pending'
            update_query = "UPDATE teacher_scheduling_preferences SET status = %s"
            # Execute the query - pass status as a tuple even if only one value
            update_cursor.execute(update_query, (new_status_value,))
            update_conn.commit()
            print(f"SCHEDULER: 已尝试更新数据库中所有教师偏好状态为 '{new_status_value}'。影响行数: {update_cursor.rowcount}") # rowcount might work

        except Exception as update_e:
            print(f"SCHEDULER: 在 finally 块中更新教师偏好状态时发生错误: {update_e}")
            if update_conn:
                try:
                    update_conn.rollback()
                    print("SCHEDULER: 教师偏好状态更新事务已回滚。")
                except Exception as rb_e:
                    print(f"SCHEDULER: 回滚教师偏好状态更新事务时发生错误: {rb_e}")
        finally:
            if update_cursor:
                try: update_cursor.close()
                except: pass
            if update_conn:
                try: update_conn.close()
                except: pass
            # --- END: Update teacher preference status ---

        return summary

# --- End of scheduler_module.py ---
